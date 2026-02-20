"""
exporters/xlsx.py

Export items to an Excel spreadsheet.
Requires: openpyxl
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = ["source", "title", "published", "url", "summary"]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_COL_WIDTHS  = {"source": 22, "title": 55, "published": 22, "url": 40, "summary": 65}


def export_xlsx(items: list[dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(items, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = item.get(col_name, "")
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_name == "url" and value:
                cell.hyperlink = value
                cell.style = "Hyperlink"
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths + freeze header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(col_name, 20)
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"  XLSX → {path}")
